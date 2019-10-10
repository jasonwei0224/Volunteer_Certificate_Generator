import openpyxl

VOLUNTEER_NUMBER = 1
VOLUNTEER_FIRST_NAME= 2
VOLUNTEER_LAST_NAME= 3
VOLUNTEER_EMAIL = 4
VOLUNTEER_HOUR= 5
VOLUNTEER_POSITION = 6
VOLUNTEER_COMMENT = 7
VOLUNTEER_CERT_LOC = 8  
class VolunteerExcelFile:
    """
    VolunteerExcelFile Class



    """


    def __init__(self, filename: str):
        """
        Constructor

        Takes in the filename or filepath
        filename: filename or path of the excel file
        workbook: the excel workbook
        worksheet: the active worksheet
        max_row: the maximum row that's been used in the excel sheet
        max_column: the maximum column that's been used in the excel sheet
        data: a dictionary that saves the data input from excel

        """

        self.__filename = filename
        self.__workbook = openpyxl.load_workbook(filename)
        self.__worksheet = self.__workbook.active
        self.__max_row = self.__worksheet.max_row
        self.__max_column = self.__worksheet.max_column
        self.__data = {}

    def get_file_name(self) -> str:
        """
        Returns the filename
        """
        return self.__filename

    def get_file_content(self) -> None:
        """
        Return the file content in a dictionary
        """
        for row in range(2, self.__max_row + 1):
            for column in range(1, self.__max_column):
                self.__data[self.__worksheet.cell(row, VOLUNTEER_FIRST_NAME).value + ' ' + self.__worksheet.cell(row, VOLUNTEER_LAST_NAME).value] = \
                    [self.__worksheet.cell(row, VOLUNTEER_NUMBER).value,
                    self.__worksheet.cell(row,VOLUNTEER_FIRST_NAME).value,
                    self.__worksheet.cell(row, VOLUNTEER_LAST_NAME).value,
                    self.__worksheet.cell(row, VOLUNTEER_EMAIL).value,
                    self.__worksheet.cell(row, VOLUNTEER_HOUR).value,
                    self.__worksheet.cell(row, VOLUNTEER_POSITION).value,
                    self.__worksheet.cell(row, VOLUNTEER_COMMENT).value]
        return self.__data

    def get_data(self) -> str:
        """
        Return the data dictionary
        """

        return self.__data
